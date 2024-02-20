
import arcpy
import openpyxl
import datetime

# Feature class location
feature_class = r'C:\Users\Nolan\Documents\ArcGIS\ArcGIS Pro Projects\DateDelete\DateCleanUpDelete\GISDBSERVER22.sde\HUD_LGIM.DBO.State_Lab_Samples'
excel_file = r"C:\Users\Nolan\Documents\ExcelSheets\CMDP_Sample_Result_Template - Copy3.xlsx"

# Open Excel file
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Get column names from Row 8 in Excel
excel_column_names = [cell.value for cell in sheet[8]]

fields_in_server = [ 'OBJECTID', 'DATE_REC_LAB', 'ROTATION', 'INSTALLDATE', 'NAME',  'ADDRESS', 'SAMPLE_DATE', 'SAMPLE_TIME', 'SAMPLEVAL', 'THRESHHIT',
                    'DISINFRULE', 'ENABLED', 'ACTIVEFLAG', 'OWNEDBY', 'OWNEDBY', 'LASTUPDATE', 'LASTEDITOR', 'TOTAL_COLIFORM_RESULTS', 'TOTAL_COLIFORM_RESULTS', 'SECTION', 
                    'SAMPLE_NUMBER', 'SAMPLE_FIXTURE', 'EMPLOYEE', 'STATIONID', 'MAP_NUMBER', 'SPDATE', 'SAMPLE_TYPE', 'LOCATION_COMMENTS', 'LOCATION_COMMENTS',   
                    'DATE_REC_LAB', 'TIME_REC_LAB', 'ANALYSIS_DATE', 'ANALYSIS_TIME', 'REC_LAB_BY', 'LOCATION_COMMENTS', 'ANALYST', 'TOTAL_COLIFORM_RESULTS', 'TESTED', 'ELEVATION', 'SAMPLE_READING', 'LOCATION_COMMENTS'  ]


fields_in_excel = excel_column_names
'''fields_in_excel = ['Sample ID*', 'Sample Received Date f', 'WS ID*', 'Facility ID*', 'Sampling Point ID*', 'Sampling Location', 'Collection Date*f', 'Collection Time (24H) f', 
'Sample Type*f', 'Sample Volume (ML) f', 'Repeat Location', 'Original Sample ID +', 'Original Reporting Lab.ID', 'Original Collection Date', 'Comment', 'Sample Collector Name', 
'Analyte*f\n[Code - Name]', None, 'A/P*f', None, 'Count', 'Units +', 'Volume (ML) +', None, 'Interference', None, 'Volume Assayed (ML) f', None, 'Method f', 'Analysis Start Date f', 
'Analysis Start Time f', 'Analysis Completed Date', 'Analysis Completed Time', 'Analyst Name', 'Analyzing Lab ID', 'Source Type', 'Comment', 'Parameter* \n[Code - Name]', None, 'Result*', 
'Result UOM*', None, 'Method', 'Analyst Name', 'Comment']'''
print(fields_in_excel)
# Create a dictionary to map Excel fields to feature class fields
field_mapping = dict(zip(fields_in_excel, fields_in_server))

# Clear existing data in the Excel sheet
for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, max_col=sheet.max_column):
    for cell in row:
        cell.value = None

# Use SearchCursor to retrieve data from the feature class
with arcpy.da.SearchCursor(feature_class, fields_in_server) as cursor:
    # Find the first empty row in the Excel sheet
    row_index = 9  # Assuming the data starts from the ninth row in Excel
    while sheet.cell(row=row_index, column=1).value is not None:
        row_index += 1

    # Write data to Excel
    for row in cursor:
        # Create a dictionary to map feature class field names to values
        data_dict = dict(zip(fields_in_server, row))

        # Write data to the next available row in Excel, treating all values as text
        for excel_field, server_field in field_mapping.items():
            sheet.cell(row=row_index, column=fields_in_excel.index(excel_field) + 1, value=str(data_dict[server_field])) # value=str(data_dict[server_field]))

        # Update the entire 'WS ID*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('WS ID*') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='0000303')

        # Update the entire 'WS ID*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Facility ID*') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='DIST.')

        # Update the entire 'WS ID*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Sample Type*f') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='Routine')

        # Update the entire 'WS ID*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Sample Volume (ML) f') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='100 mg/l')

        # Update the entire 'WS ID*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Analyte*f\n[Code - Name]') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='3100 - COLIFORM (TCR)')

        # Update the entire 'WS ID*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Analyzing Lab ID') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='03132')

        # Update the entire 'WS ID*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Parameter* \n[Code - Name]') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='0999 - Chlorine')

        # Update the entire 'WS ID*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Result UOM*') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='mg/L')
        # Copy the above code for each repeating column


        row_index += 1

# Save the modified Excel file
workbook.save(excel_file)

# Open Excel file
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Delete rows where 'Sample Received Date f' equals 'home'
delete_rows = []
for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Collection Date*f') + 1).value
    if cell_value == 'None':
        delete_rows.append(row)

# Print to check which rows are going to be deleted
print("Rows to be deleted:", delete_rows)

for row in reversed(delete_rows):
    print("Deleting row:", row)
    sheet.delete_rows(row)

# Save the final modified Excel file after deleting rows
workbook.save(excel_file)


# Open Excel file
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active


# Format 'Collection Date*f' to MM/DD/YYYY
for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Sample Received Date f') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Sample Received Date f') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")

# Save the final modified Excel file after formatting dates
workbook.save(excel_file)


# Open Excel file
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Collection Date*f') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Collection Date*f') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")

# Save the final modified Excel file after formatting dates
workbook.save(excel_file)


# Open Excel file
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Analysis Start Date f') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Analysis Start Date f') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")

# Save the final modified Excel file after formatting dates
workbook.save(excel_file)


# Open Excel file
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Analysis Completed Date') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Analysis Completed Date') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")

# Save the final modified Excel file after formatting dates
workbook.save(excel_file)


# Open Excel file
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Clear data in 'Repeat Location' and 'Original Sample ID' columns
columns_to_clear = ['Repeat Location', 'Original Sample ID +', 'Original Reporting Lab.ID', 'Original Collection Date', 'Sample Collector Name', 'Comment', 'Count', 'Units +', 'Volume (ML) +', 'Interference',
                     'Volume Assayed (ML) f', 'Method f', 'Source Type' ]

for column in columns_to_clear:
    column_index = fields_in_excel.index(column) + 1  # Adding 1 to convert from 0-based to 1-based index

    for row in range(9, sheet.max_row + 1):
        sheet.cell(row=row, column=column_index).value=None

# Save the final modified Excel file after clearing data
workbook.save(excel_file)
