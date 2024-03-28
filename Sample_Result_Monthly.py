
import arcpy
import openpyxl
import datetime

# Feature class location
feature_class = r'\\portalserver\Production Projects\GISDBSERVER22.sde\HUD_LGIM.DBO.State_Lab_Samples'
excel_file = r"\\VSERVER22\ForEveryone\Nolan\CMDP_Sample_Result_Template - Excel.xlsx" 

# Open Excel file
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Get column names from Row 8 in Excel
excel_column_names = [cell.value for cell in sheet[8]]

fields_in_server = [ 'OBJECTID', 'DATE_REC_LAB', 'ROTATION', 'INSTALLDATE', 'NAME',  'ADDRESS', 'SAMPLE_DATE', 'SAMPLE_TIME', 'SAMPLEVAL', 'THRESHHIT',
                    'DISINFRULE', 'ENABLED', 'ACTIVEFLAG', 'OWNEDBY', 'OWNEDBY', 'LASTUPDATE', 'LASTEDITOR', 'OBJECTID', 'TOTAL_COLIFORM_RESULTS', 'SECTION', 
                    'SAMPLE_NUMBER', 'SAMPLE_FIXTURE', 'EMPLOYEE', 'STATIONID', 'MAP_NUMBER', 'SPDATE', 'SAMPLE_TYPE', 'LOCATION_COMMENTS', 'LOCATION_COMMENTS',   
                    'DATE_REC_LAB', 'TIME_REC_LAB', 'ANALYSIS_DATE', 'ANALYSIS_TIME', 'REC_LAB_BY', 'LOCATION_COMMENTS', 'ANALYST', 'TOTAL_COLIFORM_RESULTS', 'TESTED', 'ELEVATION', 'SAMPLE_READING', 'LOCATION_COMMENTS'  ]


fields_in_excel = excel_column_names
'''fields_in_excel = ['Sample ID*', 'Sample Received Date f', 'WS ID*', 'Facility ID*', 'Sampling Point ID*', 'Sampling Location', 'Collection Date*f', 'Collection Time (24H) f', 
'Sample Type*f', 'Sample Volume (ML) f', 'Repeat Location', 'Original Sample ID +', 'Original Reporting Lab.ID', 'Original Collection Date', 'Comment', 'Sample Collector Name', 
'Analyte*f\n[Code - Name]', None, 'A/P*f', None, 'Count', 'Units +', 'Volume (ML) +', None, 'Interference', None, 'Volume Assayed (ML) f', None, 'Method f', 'Analysis Start Date f', 
'Analysis Start Time f', 'Analysis Completed Date', 'Analysis Completed Time', 'Analyst Name', 'Analyzing Lab ID', 'Source Type', 'Comment', 'Parameter* \n[Code - Name]', None, 'Result*', 
'Result UOM*', None, 'Method', 'Analyst Name', 'Comment']'''
print(fields_in_excel)
# Create a dictionary to map Excel fields to feature class fields
field_mapping = dict(zip(fields_in_excel, fields_in_server))

# Clear existing data in the Excel sheet
for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, max_col=sheet.max_column):
    for cell in row:
        cell.value = None

# Use SearchCursor to retrieve data from the feature class
with arcpy.da.SearchCursor(feature_class, fields_in_server) as cursor:
    # Find the first empty row in the Excel sheet
    row_index = 9  # Assuming the data starts from the ninth row in Excel
    while sheet.cell(row=row_index, column=1).value is not None:
        row_index += 1
    

    # Write data to Excel
    for row in cursor:
        # Create a dictionary to map feature class field names to values
        data_dict = dict(zip(fields_in_server, row))
        # Write data to the next available row in Excel, treating all values as text
        for excel_field, server_field in field_mapping.items():
            sheet.cell(row=row_index, column=fields_in_excel.index(excel_field) + 1, value=str(data_dict[server_field])) # value=str(data_dict[server_field]))

        # Update the entire 'WS ID*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('WS ID*') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='TN0000303')

        # Update the entire 'Facility ID*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Facility ID*') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='DIST.')

        # Update the entire 'Sample Type*f' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Sample Type*f') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='Routine')

        # Update the entire 'Sample Volume (ML) f*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Sample Volume (ML) f') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='100')

        # Update the entire 'Analyte*f\n[Code - Name]' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Analyte*f\n[Code - Name]') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='3100 - COLIFORM (TCR)')

        # Update the entire 'Analyzing Lab ID' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Analyzing Lab ID') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='03132')

        # Update the entire 'Parameter* \n[Code - Name]' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Parameter* \n[Code - Name]') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='0999 - Chlorine')

        # Update the entire 'Result UOM*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Result UOM*') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='mg/L')
        
        # Update the entire 'Result UOM*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Facility ID*') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='DS01')

        # Update the entire 'Result UOM*' column in the Excel sheet for the current row
        ws_id_column_index = fields_in_excel.index('Sampling Point ID*') + 1  # Adding 1 to convert from 0-based to 1-based index
        sheet.cell(row=row_index, column=ws_id_column_index, value='DS01')

        row_index += 1


# Delete rows where 'Collection Date*f' equals 'None'
delete_rows = []
for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Collection Date*f') + 1).value
    if cell_value == 'None':
        delete_rows.append(row)

# Print to check which rows are going to be deleted
print("Rows to be deleted:", delete_rows)

for row in reversed(delete_rows):
    print("Deleting row:", row)
    sheet.delete_rows(row)

# Save the final modified Excel file after deleting rows
# workbook.save(excel_file)


# Format 'Sample Received Date f' to MM/DD/YYYY
for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Sample Received Date f') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S.%f").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Sample Received Date f') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")


# Format 'Sample Received Date f' to MM/DD/YYYY
for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Sample Received Date f') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Sample Received Date f') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")


for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Collection Date*f') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Collection Date*f') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")


for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Collection Date*f') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S.%f").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Collection Date*f') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")


for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Analysis Start Date f') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S.%f").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Analysis Start Date f') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")


for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Analysis Start Date f') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Analysis Start Date f') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")


for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Analysis Completed Date') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S.%f").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Analysis Completed Date') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")


for row in range(9, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=fields_in_excel.index('Analysis Completed Date') + 1).value
    
    if cell_value:
        try:
            # Parse the existing date
            existing_date = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S").date()
            
            # Format the date as MM/DD/YYYY
            formatted_date = existing_date.strftime("%m/%d/%Y")
            
            # Update the cell with the formatted date
            sheet.cell(row=row, column=fields_in_excel.index('Analysis Completed Date') + 1, value=formatted_date)
        except ValueError:
            print(f"Unable to parse date in row {row}: {cell_value}")


# Clear data in 'Repeat Location' and 'Original Sample ID' columns
columns_to_clear = ['Sample ID*', 'Repeat Location', 'Original Sample ID +', 'Original Reporting Lab.ID', 'Original Collection Date', 
                    'Sample Collector Name', 'Comment', 'Count', 'Units +', 'Volume (ML) +', 'Interference','Volume Assayed (ML) f', 'Method f', 'Source Type', 'A/P*f' ]

for column in columns_to_clear:
    column_index = fields_in_excel.index(column) + 1  # Adding 1 to convert from 0-based to 1-based index

    for row in range(9, sheet.max_row + 1):
        sheet.cell(row=row, column=column_index).value=None


# Get current date
time = datetime.datetime.now()
sample_id_date = time.strftime("%m%d%y")

# Create Sample ID
# Find the first empty row in the Excel sheet
row_index = 9  # Assuming the data starts from the ninth row in Excel
while sheet.cell(row=row_index, column=1).value is not None:
    row_index += 1

# Update the 'Sample ID*' column with row numbers starting from 1 to 70
for row_number in range(row_index, min(row_index + 70, sheet.max_row + 1)):
    sample_id = f"{sample_id_date}-{row_number - row_index + 1}"
    sheet.cell(row=row_number, column=fields_in_excel.index('Sample ID*') + 1, value=sample_id)



# Find the index of the 'Comment' and 'A/P*f' columns
comment_column_index = fields_in_excel.index('Comment.') + 1  # Adding 1 to convert from 0-based to 1-based index
ap_column_index = fields_in_excel.index('A/P*f') + 1  # Adding 1 to convert from 0-based to 1-based index

# Iterate through rows and update 'A/P*f' column based on 'Comment'
for row in range(9, sheet.max_row + 1):
    comment_value = sheet.cell(row=row, column=comment_column_index).value

    # Check if 'Comment' contains 'Negative' (case-insensitive)
    if comment_value and 'negative' in comment_value.strip().lower():
        sheet.cell(row=row, column=ap_column_index, value='Absent')

            # Check if 'Comment' contains 'Positive' (case-insensitive)
    elif comment_value and 'positive' in comment_value.strip().lower():
        sheet.cell(row=row, column=ap_column_index, value='Present')


# Save the modified Excel file
workbook.save(excel_file)